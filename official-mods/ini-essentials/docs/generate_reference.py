#!/usr/bin/env python3
"""Generate the bilingual INI Essentials reference workbook from tracked CSV catalogs."""

from __future__ import annotations

import argparse
import csv
import os
import re
import warnings
import zipfile
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree
from xml.sax.saxutils import escape, quoteattr

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
FIELD_SOURCE = ROOT / "src/main/resources/ini_essentials/fields.csv"
EVENT_DATA_SOURCE = ROOT / "src/main/resources/ini_essentials/event-data.csv"
OUTPUT = Path(__file__).with_name("INI Essentials Unit Modding Reference.xlsx")

FONT_NAME = "Arial"
TITLE_COLOR = "EFEFEF"
NOTICE_COLOR = "00FF00"
INDEX_COLOR = "D9D9D9"
EVENT_INDEX_COLOR = "B2EBF2"
GRID_COLOR = "A6A6A6"
ROW_ALT_COLOR = "F3F3F3"
WHITE = "FFFFFF"
BLACK = "000000"

# Colors copied from the corresponding section bands in the original 1.15/1.16 reference.
SECTION_PALETTES = {
    "core": ("0F9D58", "0D904F"),
    "canbuild": ("8E7CC3", "674EA7"),
    "graphics": ("EA9999", "CC0000"),
    "attack": ("6D9EEB", "3C78D8"),
    "turret": ("8E7CC3", "674EA7"),
    "projectile": ("BF9000", "7F6000"),
    "movement": ("D5A6BD", "A64D79"),
    "ai": ("DD7E6B", "A61C00"),
    "leg_arm": ("134F5C", "0C343D"),
    "attachment": ("CC4125", "85200C"),
    "action": ("FF9900", "B45F06"),
    "event": ("00A6B8", "007C91"),
    # Saturated indigo stays visually distinct while preserving high white-text contrast.
    "geometry": ("5E35B1", "311B92"),
    "fog": ("455A64", "263238"),
    "math": ("C62828", "8E0000"),
    "effect": ("A64D79", "741B47"),
    "animation": ("45818E", "134F5C"),
    "resource": ("134F5C", "0C343D"),
    "comment_template": ("134F5C", "0C343D"),
    "other": ("4A86E8", "1155CC"),
}

COLUMN_WIDTHS = [15, 28, 18, 68, 42, 22, 16, 23]
SHORTCUT_COLUMNS = 4


@dataclass(frozen=True)
class ReferenceGroup:
    key: str
    section_en: str
    section_zh: str
    summary_en: str
    summary_zh: str
    palette: str
    rows: tuple[dict[str, str], ...]
    event_data: bool = False


@dataclass(frozen=True)
class NavigationButton:
    """An invisible DrawingML hyperlink laid over a styled cell range."""

    sheet_index: int
    name: str
    target: str
    first_row: int
    first_column: int
    last_row: int
    last_column: int
    equal_slot: int | None = None
    equal_slot_count: int | None = None
    text: str | None = None
    background: str | None = None


def load_rows(source: Path) -> list[dict[str, str]]:
    with source.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def section_key(section: str) -> str:
    value = section.lower()
    if "[event_" in value:
        return "event"
    if "geometry" in value:
        return "geometry"
    if "fog" in value:
        return "fog"
    if "logicboolean" in value or "numeric expression" in value:
        return "math"
    if "core" in value:
        return "core"
    if "action" in value:
        return "action"
    if "graphics" in value:
        return "graphics"
    if "attack" in value:
        return "attack"
    if "turret" in value:
        return "turret"
    if "projectile" in value:
        return "projectile"
    if "movement" in value:
        return "movement"
    if "attachment" in value:
        return "attachment"
    if "effect" in value:
        return "effect"
    if "animation" in value:
        return "animation"
    if "resource" in value:
        return "resource"
    if "canbuild" in value:
        return "canbuild"
    if "leg" in value or "arm" in value:
        return "leg_arm"
    if "comment" in value or "template" in value:
        return "comment_template"
    if "ai" in value:
        return "ai"
    return "other"


def make_groups(field_rows: list[dict[str, str]],
                event_rows: list[dict[str, str]]) -> list[ReferenceGroup]:
    grouped: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
    section_names: dict[str, str] = {}
    for row in field_rows:
        key = section_key(row["section"])
        grouped.setdefault(key, []).append(row)
        section_names.setdefault(key, row["section"])

    groups: list[ReferenceGroup] = []
    for key, rows in grouped.items():
        if key == "core":
            section_en, section_zh = "[core]", "[core]"
            summary_en = "Core unit functions and opt-in static unit fields"
            summary_zh = "单位核心功能与可选的静态单位字段"
        elif key == "action":
            section_en = "[action_NAME] / [hiddenAction_NAME]"
            section_zh = "[action_NAME] / [hiddenAction_NAME]"
            summary_en = "Action effects that run through the native custom-action chain"
            summary_zh = "通过原版自定义动作链执行的动作效果"
        elif key == "event":
            section_en, section_zh = "[event_NAME]", "[event_NAME]"
            summary_en = "Ordered queued-event rules that leave the native action-effect order intact"
            summary_zh = "不改变原版动作效果顺序的有序队列事件规则"
        elif key == "geometry":
            section_en, section_zh = "[geometry_NAME]", "[geometry_NAME]"
            summary_en = "Reusable runtime geometry masks for fog and future gameplay consumers"
            summary_zh = "供迷雾及后续玩法功能复用的运行时几何遮罩"
        elif key == "fog":
            section_en, section_zh = "[fog_NAME]", "[fog_NAME]"
            summary_en = "Reusable per-team fog operations driven by geometry masks"
            summary_zh = "由几何遮罩驱动、可复用的分队迷雾操作"
        elif key == "math":
            section_en = "Runtime LogicBoolean numeric expressions"
            section_zh = "运行时 LogicBoolean 数值表达式"
            summary_en = "Additional deterministic numeric functions available in dynamic INI expressions"
            summary_zh = "动态 INI 表达式中可用的额外确定性数值函数"
        else:
            section_en = section_names[key]
            section_zh = section_names[key]
            summary_en = "INI Essentials additions for this native section"
            summary_zh = "该原版节对应的 INI Essentials 扩展"
        groups.append(ReferenceGroup(
            key, section_en, section_zh, summary_en, summary_zh,
            key, tuple(rows), False))

    events: OrderedDict[str, list[dict[str, str]]] = OrderedDict()
    for row in event_rows:
        events.setdefault(row["event"], []).append(row)
    for event, rows in events.items():
        event_key = re.sub(r"[^a-z0-9]+", "_", event.lower()).strip("_")
        groups.append(ReferenceGroup(
            f"event_{event_key}",
            f"{event} — eventData(...)",
            f"{event} — eventData(...) 事件数据",
            f"Extra context values for the native {event} action event",
            f"为原版 {event} 动作事件补充的上下文值",
            "event", tuple(rows), True))
    return groups


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb(color))


def argb(color: str) -> str:
    return color if len(color) == 8 else "FF" + color


def define_target(workbook: Workbook, name: str,
                  sheet_title: str, coordinate: str) -> None:
    escaped_title = sheet_title.replace("'", "''")
    workbook.defined_names.add(DefinedName(
        name, attr_text=f"'{escaped_title}'!${coordinate[0]}${coordinate[1:]}"))


def add_navigation_button(buttons: list[NavigationButton], sheet,
                          name: str, target: str,
                          first_row: int, first_column: int,
                          last_row: int, last_column: int, *,
                          equal_slot: int | None = None,
                          equal_slot_count: int | None = None,
                          text: str | None = None,
                          background: str | None = None) -> None:
    buttons.append(NavigationButton(
        sheet.parent.index(sheet) + 1,
        name,
        target,
        first_row,
        first_column,
        last_row,
        last_column,
        equal_slot,
        equal_slot_count,
        text,
        background,
    ))


def style_range(sheet, row: int, start: int, end: int, *,
                background: str, font_color: str = BLACK, bold: bool = False,
                horizontal: str = "left", size: float = 10.0) -> None:
    edge = Side(style="thin", color=argb(GRID_COLOR))
    for column in range(start, end + 1):
        cell = sheet.cell(row, column)
        cell.fill = fill(background)
        cell.font = Font(
            name=FONT_NAME, size=size, bold=bold, color=argb(font_color))
        cell.alignment = Alignment(
            horizontal=horizontal, vertical="center", wrap_text=True)
        cell.border = Border(left=edge, right=edge, top=edge, bottom=edge)


def shortcut_rows(group_count: int) -> int:
    return max(1, (group_count + SHORTCUT_COLUMNS - 1) // SHORTCUT_COLUMNS)


def add_shortcuts(sheet, groups: list[ReferenceGroup], starts: dict[str, int],
                  start_row: int, chinese: bool,
                  buttons: list[NavigationButton]) -> int:
    spans = ((1, 2), (3, 4), (5, 6), (7, 8))
    language = "zh" if chinese else "en"
    for index, group in enumerate(groups):
        row = start_row + index // SHORTCUT_COLUMNS
        slot = index % SHORTCUT_COLUMNS
        first, last = spans[slot]
        sheet.merge_cells(start_row=row, start_column=first,
                          end_row=row, end_column=last)
        cell = sheet.cell(row, first)
        cell.value = group.section_zh if chinese else group.section_en
        target_name = f"rf_{language}_{group.key}"
        define_target(sheet.parent, target_name, sheet.title,
                      f"A{starts[group.key]}")
        section_color, _ = SECTION_PALETTES.get(
            group.palette, SECTION_PALETTES["other"])
        add_navigation_button(
            buttons, sheet, f"shortcut_{language}_{group.key}", target_name,
            row, 1, row, 8,
            equal_slot=slot, equal_slot_count=SHORTCUT_COLUMNS,
            text=cell.value, background=section_color)
        style_range(sheet, row, first, last,
                    background=section_color, font_color=WHITE,
                    bold=True, horizontal="center", size=10.0)
        cell.font = Font(
            name=FONT_NAME, size=10.0, bold=True,
            color=argb(WHITE))
        sheet.row_dimensions[row].height = 27
    return start_row + shortcut_rows(len(groups))


def add_group(sheet, group: ReferenceGroup, start_row: int,
              chinese: bool, shortcut_target: str,
              buttons: list[NavigationButton]) -> int:
    section_color, header_color = SECTION_PALETTES.get(
        group.palette, SECTION_PALETTES["other"])
    section_name = group.section_zh if chinese else group.section_en
    summary = group.summary_zh if chinese else group.summary_en

    style_range(sheet, start_row, 1, 8,
                background=section_color, font_color=WHITE, size=10.0)
    sheet.cell(start_row, 1, "Version" if not chinese else "版本")
    sheet.cell(start_row, 2, "Section" if not chinese else "节")
    sheet.cell(start_row, 4, section_name)
    sheet.cell(start_row, 5, summary)
    sheet.merge_cells(start_row=start_row, start_column=5,
                      end_row=start_row, end_column=7)
    back = sheet.cell(start_row, 8, "↑ Shortcuts" if not chinese else "↑ 返回节导航")
    language = "zh" if chinese else "en"
    add_navigation_button(
        buttons, sheet, f"back_{language}_{group.key}", shortcut_target,
        start_row, 8, start_row, 8)
    back.font = Font(
        name=FONT_NAME, size=10.0, bold=True,
        color=argb(WHITE))
    sheet.cell(start_row, 4).font = Font(
        name=FONT_NAME, size=10.0, bold=True, color=argb(WHITE))
    sheet.row_dimensions[start_row].height = 34

    header_row = start_row + 1
    headers = (
        ["加入版本", "代码", "值类型", "说明与用法", "实际用法与示例",
         "扩展类型", "默认值", "联机影响"]
        if chinese else
        ["Version Added", "Code", "Value Type", "Description and usage",
         "Actual usage with examples", "Extension Type", "Default",
         "Multiplayer Impact"]
    )
    style_range(sheet, header_row, 1, 8,
                background=header_color, font_color=WHITE,
                bold=True, horizontal="center", size=11.0)
    for column, value in enumerate(headers, 1):
        sheet.cell(header_row, column, value)
    sheet.row_dimensions[header_row].height = 31

    description_key = "description_zh" if chinese else "description_en"
    for offset, row in enumerate(group.rows, 2):
        target_row = start_row + offset
        if group.event_data:
            values = [
                row["version_added"], row["event_data_name"], row["value_type"],
                row[description_key], row["example"],
                "native_event_data", "available during event", row["multiplayer_impact"],
            ]
        else:
            values = [
                row["version_added"], row["code"], row["value_type"],
                row[description_key], row["example"], row["extension_type"],
                row["default"], row["multiplayer_impact"],
            ]
        background = WHITE if offset % 2 == 0 else ROW_ALT_COLOR
        style_range(sheet, target_row, 1, 8, background=background)
        for column, value in enumerate(values, 1):
            cell = sheet.cell(target_row, column, value)
            cell.alignment = Alignment(
                horizontal="center" if column in (1, 3, 6, 7, 8) else "left",
                vertical="top", wrap_text=True)
            if column == 2:
                cell.font = Font(
                    name=FONT_NAME, size=10.0, bold=True, color=argb(BLACK))
        length = max(len(str(values[3])), len(str(values[4])))
        sheet.row_dimensions[target_row].height = max(28, min(88, 19 + length // 55 * 13))

    return start_row + len(group.rows) + 3


def add_reference_sheet(workbook: Workbook, title: str,
                        groups: list[ReferenceGroup], chinese: bool,
                        buttons: list[NavigationButton]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:H1")
    sheet["A1"] = ("INI Essentials Unit Modding Reference"
                   if not chinese else "INI Essentials 单位模组代码表")
    style_range(sheet, 1, 1, 8,
                background=TITLE_COLOR, bold=True, horizontal="center", size=14.0)
    sheet.row_dimensions[1].height = 38

    sheet.merge_cells("A2:H2")
    sheet["A2"] = (
        "Generated from tracked CSV catalogs. Section colors follow the original Rusted Warfare reference."
        if not chinese else
        "由仓库内 CSV 代码表生成；各节颜色沿用原版 Rusted Warfare 代码表。"
    )
    style_range(sheet, 2, 1, 8,
                background=NOTICE_COLOR, bold=True, horizontal="center", size=10.0)
    sheet.row_dimensions[2].height = 29

    sheet.merge_cells("A3:H3")
    sheet["A3"] = "Section shortcuts — click to jump" if not chinese else "节快捷导航——点击即可跳转"
    style_range(sheet, 3, 1, 8,
                background=INDEX_COLOR, bold=True, horizontal="center", size=11.0)
    sheet.row_dimensions[3].height = 28

    regular_groups = [group for group in groups if not group.event_data]
    event_groups = [group for group in groups if group.event_data]
    index_rows = shortcut_rows(len(regular_groups))
    event_category_row = 4 + index_rows
    first_group_row = event_category_row + 2
    regular_starts: dict[str, int] = {}
    cursor = first_group_row
    for group in regular_groups:
        regular_starts[group.key] = cursor
        cursor += len(group.rows) + 3

    language = "zh" if chinese else "en"
    shortcut_target = f"rf_{language}_shortcuts"
    define_target(workbook, shortcut_target, sheet.title, "A3")
    define_target(workbook, f"rf_{language}_title", sheet.title, "A1")
    add_shortcuts(sheet, regular_groups, regular_starts, 4, chinese, buttons)

    event_category_target = f"rf_{language}_event_data"
    define_target(workbook, event_category_target, sheet.title, f"A{cursor}")
    sheet.merge_cells(start_row=event_category_row, start_column=1,
                      end_row=event_category_row, end_column=8)
    category = sheet.cell(
        event_category_row, 1,
        "Native eventData extensions ↓" if not chinese else "原版事件数据扩展 ↓")
    style_range(sheet, event_category_row, 1, 8,
                background=SECTION_PALETTES["event"][0], font_color=WHITE,
                bold=True, horizontal="center", size=11.0)
    category.font = Font(
        name=FONT_NAME, size=11.0, bold=True, color=argb(WHITE))
    sheet.row_dimensions[event_category_row].height = 29
    add_navigation_button(
        buttons, sheet, f"open_{language}_event_data", event_category_target,
        event_category_row, 1, event_category_row, 8)

    cursor = first_group_row
    for group in regular_groups:
        cursor = add_group(
            sheet, group, cursor, chinese, shortcut_target, buttons)

    event_banner_row = cursor
    sheet.merge_cells(start_row=event_banner_row, start_column=1,
                      end_row=event_banner_row, end_column=8)
    sheet.cell(event_banner_row, 1, (
        "Native autoTriggerOnEvent data extensions"
        if not chinese else "原版 autoTriggerOnEvent 事件数据扩展"))
    style_range(sheet, event_banner_row, 1, 8,
                background=SECTION_PALETTES["event"][1], font_color=WHITE,
                bold=True, horizontal="center", size=14.0)
    sheet.row_dimensions[event_banner_row].height = 42

    event_index_row = event_banner_row + 1
    event_shortcut_target = f"rf_{language}_event_shortcuts"
    define_target(workbook, event_shortcut_target, sheet.title,
                  f"A{event_index_row}")
    sheet.merge_cells(start_row=event_index_row, start_column=1,
                      end_row=event_index_row, end_column=8)
    sheet.cell(event_index_row, 1, (
        "Event shortcuts — click to jump"
        if not chinese else "事件快捷导航——点击即可跳转"))
    style_range(sheet, event_index_row, 1, 8,
                background=EVENT_INDEX_COLOR, bold=True,
                horizontal="center", size=11.0)
    sheet.row_dimensions[event_index_row].height = 28

    event_shortcut_row = event_index_row + 1
    event_first_group_row = (
        event_shortcut_row + shortcut_rows(len(event_groups)) + 1)
    event_starts: dict[str, int] = {}
    event_cursor = event_first_group_row
    for group in event_groups:
        event_starts[group.key] = event_cursor
        event_cursor += len(group.rows) + 3
    add_shortcuts(
        sheet, event_groups, event_starts, event_shortcut_row, chinese, buttons)
    event_cursor = event_first_group_row
    for group in event_groups:
        event_cursor = add_group(
            sheet, group, event_cursor, chinese, event_shortcut_target, buttons)

    for index, width in enumerate(COLUMN_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = f"A{first_group_row}"
    sheet.auto_filter.ref = None
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True


def add_about_sheet(workbook: Workbook,
                    buttons: list[NavigationButton]) -> None:
    sheet = workbook.active
    sheet.title = "About 关于"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "INI Essentials Unit Modding Reference"
    style_range(sheet, 1, 1, 8,
                background=TITLE_COLOR, bold=True, horizontal="center", size=16.0)
    sheet.row_dimensions[1].height = 42

    lines = [
        "Generated from fields.csv and event-data.csv. Do not edit this workbook as the source.",
        "本表由 fields.csv 和 event-data.csv 自动生成，请勿把工作簿作为源文件直接修改。",
        "Native keys with native-valid values stay on the native parser path.",
        "原版字段使用原版合法值时，始终保留原版解析路径。",
        "Extensions activate only for a documented new key, value range, format, or event-data name.",
        "只有代码表明确记录的新字段、新取值范围、新格式或事件数据名才会激活扩展。",
    ]
    for row, value in enumerate(lines, 3):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        sheet.cell(row, 1, value)
        style_range(sheet, row, 1, 8,
                    background=WHITE if row % 2 else ROW_ALT_COLOR, size=11.0)
        sheet.row_dimensions[row].height = 27

    for row, (label, target, color) in enumerate((
            ("Open English reference / 打开英文代码表", "English", "4A86E8"),
            ("打开简体中文代码表 / Open Chinese reference", "简体中文", "0F9D58")), 10):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        cell = sheet.cell(row, 2, label)
        target_name = "rf_en_title" if target == "English" else "rf_zh_title"
        add_navigation_button(
            buttons, sheet, f"open_{target_name}", target_name,
            row, 2, row, 7)
        style_range(sheet, row, 2, 7,
                    background=color, font_color=WHITE,
                    bold=True, horizontal="center", size=11.0)
        cell.font = Font(
            name=FONT_NAME, size=11.0, bold=True,
            color=argb(WHITE))
        sheet.row_dimensions[row].height = 31

    for index, width in enumerate(COLUMN_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook() -> tuple[Workbook, list[NavigationButton]]:
    groups = make_groups(load_rows(FIELD_SOURCE), load_rows(EVENT_DATA_SOURCE))
    workbook = Workbook()
    buttons: list[NavigationButton] = []
    workbook.properties.title = "INI Essentials Unit Modding Reference"
    workbook.properties.creator = "Rusted Fabric Loader / INI Essentials"
    workbook.properties.description = "Generated bilingual reference for INI Essentials"
    add_about_sheet(workbook, buttons)
    add_reference_sheet(workbook, "English", groups, False, buttons)
    add_reference_sheet(workbook, "简体中文", groups, True, buttons)
    return workbook, buttons


DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
DRAWING_REL_TYPE = OFFICE_REL_NS + "/drawing"
HYPERLINK_REL_TYPE = OFFICE_REL_NS + "/hyperlink"
DRAWING_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.drawing+xml")


def column_pixels(sheet, column: int) -> int:
    width = sheet.column_dimensions[get_column_letter(column)].width
    if width is None:
        width = sheet.sheet_format.defaultColWidth or 8.43
    return max(1, int(width * 7 + 5))


def row_pixels(sheet, row: int) -> int:
    height = sheet.row_dimensions[row].height
    if height is None:
        height = sheet.sheet_format.defaultRowHeight or 15
    return max(1, int(points_to_pixels(height)))


def column_marker(sheet, pixel_offset: int) -> tuple[int, int]:
    """Convert an absolute horizontal pixel offset to a DrawingML cell marker."""
    remaining = max(0, pixel_offset)
    column = 1
    while remaining >= column_pixels(sheet, column):
        remaining -= column_pixels(sheet, column)
        column += 1
    return column - 1, pixels_to_EMU(remaining)


def shape_xml(button: NavigationButton, sheet, shape_id: int,
              relationship_id: str) -> str:
    first_row = button.first_row - 1
    last_row = button.last_row
    area_x = sum(
        column_pixels(sheet, column)
        for column in range(1, button.first_column))
    y = pixels_to_EMU(sum(
        row_pixels(sheet, row)
        for row in range(1, button.first_row)))
    area_width = sum(
        column_pixels(sheet, column)
        for column in range(button.first_column, button.last_column + 1))
    if button.equal_slot is not None:
        if not button.equal_slot_count or not 0 <= button.equal_slot < button.equal_slot_count:
            raise ValueError(f"invalid equal navigation slot for {button.name}")
        left = area_x + area_width * button.equal_slot // button.equal_slot_count
        right = area_x + area_width * (button.equal_slot + 1) // button.equal_slot_count
    else:
        left = area_x
        right = area_x + area_width
    first_column, first_column_offset = column_marker(sheet, left)
    last_column, last_column_offset = column_marker(sheet, right)
    x = pixels_to_EMU(left)
    width = pixels_to_EMU(right - left)
    height = pixels_to_EMU(sum(
        row_pixels(sheet, row)
        for row in range(button.first_row, button.last_row + 1)))
    if button.background is None:
        shape_style = '<a:noFill/><a:ln><a:noFill/></a:ln>'
        text_body = (
            '<xdr:txBody><a:bodyPr vertOverflow="clip" horzOverflow="clip" '
            'wrap="none" lIns="0" tIns="0" rIns="0" bIns="0" anchor="ctr"/>'
            '<a:lstStyle/><a:p/></xdr:txBody>')
    else:
        shape_style = (
            f'<a:solidFill><a:srgbClr val="{button.background}"/></a:solidFill>'
            f'<a:ln w="9525"><a:solidFill><a:srgbClr val="{GRID_COLOR}"/>'
            '</a:solidFill></a:ln>')
        label = escape(button.text or "")
        text_body = (
            '<xdr:txBody><a:bodyPr vertOverflow="clip" horzOverflow="clip" '
            'wrap="square" lIns="45720" tIns="0" rIns="45720" bIns="0" anchor="ctr"/>'
            '<a:lstStyle/><a:p><a:pPr algn="ctr"/>'
            '<a:r><a:rPr lang="en-US" sz="1000" b="1">'
            f'<a:solidFill><a:srgbClr val="{WHITE}"/></a:solidFill>'
            f'<a:latin typeface="{FONT_NAME}"/></a:rPr><a:t>{label}</a:t></a:r>'
            '<a:endParaRPr lang="en-US" sz="1000"/></a:p></xdr:txBody>')
    return (
        '<xdr:twoCellAnchor editAs="oneCell">'
        '<xdr:from>'
        f'<xdr:col>{first_column}</xdr:col><xdr:colOff>{first_column_offset}</xdr:colOff>'
        f'<xdr:row>{first_row}</xdr:row><xdr:rowOff>0</xdr:rowOff>'
        '</xdr:from>'
        '<xdr:to>'
        f'<xdr:col>{last_column}</xdr:col><xdr:colOff>{last_column_offset}</xdr:colOff>'
        f'<xdr:row>{last_row}</xdr:row><xdr:rowOff>0</xdr:rowOff>'
        '</xdr:to>'
        '<xdr:sp macro="" textlink=""><xdr:nvSpPr>'
        f'<xdr:cNvPr id="{shape_id}" name={quoteattr(button.name)}>'
        f'<a:hlinkClick xmlns:r="{OFFICE_REL_NS}" r:id="{relationship_id}"/>'
        '</xdr:cNvPr><xdr:cNvSpPr/></xdr:nvSpPr>'
        '<xdr:spPr>'
        f'<a:xfrm><a:off x="{x}" y="{y}"/>'
        f'<a:ext cx="{width}" cy="{height}"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        f'{shape_style}'
        '</xdr:spPr>'
        f'{text_body}</xdr:sp><xdr:clientData/>'
        '</xdr:twoCellAnchor>'
    )


def drawing_xml(sheet, buttons: list[NavigationButton]) -> bytes:
    shapes = ''.join(
        shape_xml(button, sheet, index + 1, f"rId{index + 1}")
        for index, button in enumerate(buttons))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{DRAWING_NS}" xmlns:a="{DRAWING_MAIN_NS}">'
        f'{shapes}</xdr:wsDr>'
    ).encode("utf-8")


def drawing_relationships_xml(buttons: list[NavigationButton]) -> bytes:
    relationships = ''.join(
        f'<Relationship Id="rId{index + 1}" Type="{HYPERLINK_REL_TYPE}" '
        f'Target={quoteattr("#" + button.target)}/>'
        for index, button in enumerate(buttons))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{REL_NS}">{relationships}</Relationships>'
    ).encode("utf-8")


def append_before(xml: bytes, closing_tag: bytes, content: bytes) -> bytes:
    position = xml.rfind(closing_tag)
    if position < 0:
        raise RuntimeError(f"Malformed XLSX XML: missing {closing_tag!r}")
    return xml[:position] + content + xml[position:]


def inject_navigation_shapes(path: Path, workbook: Workbook,
                             buttons: list[NavigationButton]) -> None:
    """Add Excel-style click overlays that open links without selecting cells."""
    grouped: OrderedDict[int, list[NavigationButton]] = OrderedDict()
    for button in buttons:
        grouped.setdefault(button.sheet_index, []).append(button)

    with zipfile.ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    content_types = entries["[Content_Types].xml"]
    for sheet_index in grouped:
        drawing_name = f"drawing{sheet_index}.xml"
        override = (
            f'<Override PartName="/xl/drawings/{drawing_name}" '
            f'ContentType="{DRAWING_CONTENT_TYPE}"/>').encode("utf-8")
        if override not in content_types:
            content_types = append_before(
                content_types, b"</Types>", override)

        worksheet_name = f"xl/worksheets/sheet{sheet_index}.xml"
        drawing_tag = (
            f'<drawing xmlns:r="{OFFICE_REL_NS}" '
            'r:id="rIdRFNavigation"/>').encode("utf-8")
        entries[worksheet_name] = append_before(
            entries[worksheet_name], b"</worksheet>", drawing_tag)

        worksheet_rels_name = (
            f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels")
        drawing_relationship = (
            f'<Relationship Id="rIdRFNavigation" Type="{DRAWING_REL_TYPE}" '
            f'Target="../drawings/{drawing_name}"/>').encode("utf-8")
        if worksheet_rels_name in entries:
            entries[worksheet_rels_name] = append_before(
                entries[worksheet_rels_name],
                b"</Relationships>", drawing_relationship)
        else:
            entries[worksheet_rels_name] = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                f'<Relationships xmlns="{REL_NS}">').encode("utf-8") \
                + drawing_relationship + b"</Relationships>"

        drawing_path = f"xl/drawings/{drawing_name}"
        entries[drawing_path] = drawing_xml(
            workbook.worksheets[sheet_index - 1], grouped[sheet_index])
        entries[f"xl/drawings/_rels/{drawing_name}.rels"] = (
            drawing_relationships_xml(grouped[sheet_index]))
    entries["[Content_Types].xml"] = content_types

    temporary = path.with_name(path.name + ".tmp")
    with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
        for name, data in entries.items():
            target.writestr(name, data)
    os.replace(temporary, path)


def navigation_signature(path: Path) -> tuple:
    result = []
    with zipfile.ZipFile(path, "r") as archive:
        drawing_names = sorted(
            name for name in archive.namelist()
            if name.startswith("xl/drawings/drawing") and name.endswith(".xml"))
        for drawing_name in drawing_names:
            rels_name = drawing_name.replace(
                "xl/drawings/", "xl/drawings/_rels/") + ".rels"
            rels_root = ElementTree.fromstring(archive.read(rels_name))
            targets = {
                relationship.attrib["Id"]: relationship.attrib["Target"]
                for relationship in rels_root
                if relationship.attrib.get("Type") == HYPERLINK_REL_TYPE
            }
            drawing_root = ElementTree.fromstring(archive.read(drawing_name))
            for properties in drawing_root.iter(
                    f"{{{DRAWING_NS}}}cNvPr"):
                hyperlink = properties.find(
                    f"{{{DRAWING_MAIN_NS}}}hlinkClick")
                if hyperlink is None:
                    continue
                relationship_id = hyperlink.attrib[
                    f"{{{OFFICE_REL_NS}}}id"]
                result.append((
                    drawing_name,
                    properties.attrib["name"],
                    targets[relationship_id],
                ))
    return tuple(sorted(result))


def expected_navigation_signature(
        buttons: list[NavigationButton]) -> tuple:
    return tuple(sorted(
        (f"xl/drawings/drawing{button.sheet_index}.xml",
         button.name, "#" + button.target)
        for button in buttons))


def workbook_signature(workbook: Workbook) -> tuple:
    result = []
    for sheet in workbook.worksheets:
        cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None and cell.hyperlink is None:
                    continue
                color = cell.fill.fgColor
                fill_value = color.rgb if color.type == "rgb" else str(color.indexed)
                location = cell.hyperlink.location if cell.hyperlink else None
                cells.append((cell.coordinate, cell.value, fill_value,
                              cell.font.bold, cell.font.color.rgb
                              if cell.font.color and cell.font.color.type == "rgb" else None,
                              location))
        result.append((sheet.title, sheet.freeze_panes,
                       tuple(sorted(str(value) for value in sheet.merged_cells.ranges)),
                       tuple(cells)))
    return tuple(result)


def defined_name_signature(workbook: Workbook) -> tuple:
    return tuple(sorted(
        (name, definition.attr_text)
        for name, definition in workbook.defined_names.items()
        if name.startswith("rf_")
    ))


def check_output(expected: Workbook,
                 buttons: list[NavigationButton]) -> None:
    if not OUTPUT.exists():
        raise SystemExit(f"Missing generated workbook: {OUTPUT}")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        actual = load_workbook(OUTPUT, read_only=False, data_only=False)
    if (workbook_signature(expected) != workbook_signature(actual)
            or defined_name_signature(expected) != defined_name_signature(actual)
            or navigation_signature(OUTPUT)
            != expected_navigation_signature(buttons)):
        raise SystemExit(
            "Generated workbook is stale; run docs/generate_reference.py and commit the result")
    print(f"Reference workbook is current: {OUTPUT}")


def main(argv: Iterable[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check", action="store_true",
        help="verify that the committed workbook matches the CSV catalogs and generator",
    )
    args = parser.parse_args(argv)
    workbook, buttons = build_workbook()
    if args.check:
        check_output(workbook, buttons)
        return
    workbook.save(OUTPUT)
    inject_navigation_shapes(OUTPUT, workbook, buttons)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
